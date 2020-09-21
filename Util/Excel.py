import os
from openpyxl import Workbook
from openpyxl import load_workbook


class Excel():

    def __init__(self):
        self._isinstance = None
        self.title = None
        self.sheet = None
        self.table = None

    def _check_title(self):
        sheet = self._isinstance[self.sheet]
        if sheet:
            self.table = sheet
        else:
            self.table = self._isinstance.create_sheet(self.sheet)
        row = self.table._current_row
        if row>0:
            for index,value in enumerate(self.title):
                assert self.table.cell(row=1, column=index+1).value == value ,'please check title name'

    def open(self, title: list, path=None, sheet='Sheet1'):
        self.title = title
        self.sheet = sheet
        if os.path.exists(path):
            self._isinstance = load_workbook(path)
            self._check_title()
            return self
        else:
            self._isinstance = Workbook()
            self.table = self._isinstance.create_sheet(self.sheet)
            return self

    def write_added_one(self, data: dict,):
        row = self.table._current_row
        if row > 0:
            for index, value in enumerate(self.title):
                self.table.cell(row + 1, index + 1, data[value])
        else:
            for index, value in enumerate(self.title):
                self.table.cell(1, index + 1, value)
            for index, value in enumerate(self.title):
                self.table.cell(row + 2, index + 1, data[value])
        return self

    def read(self):
        pass

    def save(self, path):
        self._isinstance.save(path)


if __name__ == '__main__':
    from Models.DC03.MeetUpModel1 import MeetUpModel

    path = '/Users/admin/PycharmProjects/Longhash_spider/Spider/MeetUp/test2.xlsx'
    data = MeetUpModel().select(
        'id',
        'name',
        'created',
        'local_time',
    ).where('created', '>=', 1546272000000).take(1).get()

    wb = Excel().open(['id', 'name', 'created', 'local_time'], path=path, sheet='Sheet2')
    for i in range(1,100):
        wb.write_added_one(data[0]).save(path=path)
